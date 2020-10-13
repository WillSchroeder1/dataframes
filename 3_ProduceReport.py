# Create a summary report of produce and total sold as shown below:


import openpyxl as xl
import pandas as pd

wb = xl.load_workbook('3-produceSales.xlsx')


ws = wb.active

maxC = ws.max_column
maxR = ws.max_row

produce_dict = {}

##print("Produce","\t","Cost Per Pound","\t","Amt Sold","\t","Total")
    
for currentrow in ws.iter_rows(min_row = 2, max_row=maxR , max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = round(float(currentrow[2].value),2)
    total = round(float(currentrow[3].value),2)

    produce_dict.setdefault(name,[0.00,0.00,0.00])

    produce_dict[name][0] = round(cost,2)

    
    produce_dict[name][1] += round(amt_sold,2)

    produce_dict[name][2] += round(total,2)



    #print(produce_dict)
    #input()

for k,v in produce_dict.items():
    produce_dict[k] = [v[0],round(v[1]),round(v[2])]

#print(produce_dict)


df = pd.DataFrame.from_dict(produce_dict)
df.index = ['Cost Per Pound', 'Quantity Sold', 'Total']
df2 = df.T.sort_values(by ='Total', ascending = False)

print('Highest total Sales:\n', df2.iloc[0])
print()
print('Lowest total Sales:\n', df2.iloc[39])
print()
print(df2.loc[['Orange' , 'Beets'],['Quantity Sold', 'Total']])
print()
print('Total Sales for Cucumbers:', df2.at['Cucumber', 'Total'])
print()
print()

df3 = df2.loc[(df2['Quantity Sold'] >= 11500) & (df2['Quantity Sold'] <= 12000)]
print(df3)
print()
print()


df4 = df3.T
print(df4)



#output to the screen

##for key in produce_dict:
##    
##    print(key,"\t\t",
##          '$',format(produce_dict[key]['cost'],',.2f'),"\t",
##          format(produce_dict[key]['amt_sold'],',.2f'),"\t",
##          '$',format(produce_dict[key]['total'],',.2f'))
##    
##
###write it to a .csv file instead
##
##    outfile = open("ProduceReport.csv", 'w')
##
##    outfile.write("Produce,Cost Per Pound,Amt Sold,Total\n")
##

##for key in produce_dict:
##        outfile.write(key + ',' +
##                      str(produce_dict[key]['cost']) + ',' +
##                      format(produce_dict[key]['amt_sold'],',.2f') + ',' +
##                      format(produce_dict[key]['total'],',.2f') + '\n'
##                      )

# why would we not want to do the above?
# becos format with ',' will create extra commas and cause problems reading
# the file

##for key in produce_dict:
##        outfile.write(key + ',' +
##                      str(produce_dict[key]['cost']) + ',' +
##                      str(round(produce_dict[key]['amt_sold'],2)) + ',' +
##                      str(round(produce_dict[key]['total'],2)) + '\n'
##                      )
##       
##        
##outfile.close()

    


    

    

    
    
